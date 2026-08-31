from decimal import Decimal
import random
import calendar
import csv
from .models import StoreReturnPolicy, StoreReturnRequest, StoreOrder
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db.models import Q
from django.contrib import messages
from django.contrib.auth import get_user_model, login
from openpyxl import Workbook
from .models import (
    Category, Product, ProductImage, ProductVideo, 
    StoreOrder, StoreOrderItem, StoreInvoice, StoreRating, 
    Wishlist, BrowsingHistory, ActivityAuditLog, CompanyPolicy, 
    Notification, CompanyBankAccount, UserAddressBook, PromoTheme, 
    StoreGlobalSetting, TaxRateProposal, StoreReturnRequest, RefundReason
)

User = get_user_model()

# ==========================================
# CUSTOMER FRONTEND & SHOPPING VIEWS
# ==========================================
from django.db.models import Q

from django.shortcuts import render, redirect
from django.db.models import Q
from .models import Category, Product, UserSearchHistory

def store_home_view(request):
    """The main storefront that logs searches and saves history."""
    categories = Category.objects.all()
    selected_category_id = request.GET.get('category')
    search_query = request.GET.get('q', '').strip()
    image_search_file = request.FILES.get('image_search')
    
    if not request.session.session_key:
        request.session.create()
    session_key = request.session.session_key

    products = Product.objects.filter(is_active=True).order_by('-id')
    
    if selected_category_id:
        products = products.filter(category_id=selected_category_id)
        
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
        # Log search history
        UserSearchHistory.objects.create(
            user=request.user if request.user.is_authenticated else None,
            session_key=session_key if not request.user.is_authenticated else None,
            keyword=search_query
        )
        
    elif image_search_file:
        filename = image_search_file.name.lower()
        for char in ['-', '_', '+', '.', 'jfif', 'jpg', 'jpeg', 'png']:
            filename = filename.replace(char, ' ')
        keywords = [w for w in filename.split() if len(w) > 2]
        if keywords:
            image_q = Q()
            for kw in keywords:
                image_q |= Q(name__icontains=kw) | Q(description__icontains=kw) | Q(visual_search_tag__icontains=kw)
            products = products.filter(image_q)

    active_promo_theme = PromoTheme.objects.filter(is_active=True).first()
    
    context = {
        'categories': categories,
        'products': products,
        'selected_category': selected_category_id,
        'search_query': search_query,
        'active_promo_theme': active_promo_theme,
    }
    
    return render(request, 'store/home.html', context)


def store_user_history_view(request):
    """Displays user search history and previously added or viewed items."""
    if not request.session.session_key:
        request.session.create()
    session_key = request.session.session_key

    if request.user.is_authenticated:
        searches = UserSearchHistory.objects.filter(user=request.user).order_by('-created_at')[:20]
    else:
        searches = UserSearchHistory.objects.filter(session_key=session_key).order_by('-created_at')[:20]

    context = {
        'searches': searches,
    }
    return render(request, 'store/history.html', context)

def product_detail_view(request, product_slug):
    product = get_object_or_404(Product, slug=product_slug, is_active=True)
    ratings = product.ratings.all().order_by('-created_at')
    
    additional_images = product.additional_images.all()
    product_videos = product.product_videos.all()

    active_promo_theme = PromoTheme.objects.filter(is_active=True).first()

    if request.user.is_authenticated:
        BrowsingHistory.objects.update_or_create(
            customer=request.user,
            product=product,
            defaults={'viewed_at': timezone.now()}
        )
        
    is_wishlisted = False
    if request.user.is_authenticated:
        is_wishlisted = Wishlist.objects.filter(customer=request.user, product=product).exists()
        
    context = {
        'product': product,
        'ratings': ratings,
        'is_wishlisted': is_wishlisted,
        'additional_images': additional_images,
        'product_videos': product_videos,
        'active_promo_theme': active_promo_theme,
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/product_detail.html', context)

    return render(request, 'store/product_detail.html', context)


@login_required
def toggle_wishlist_view(request, product_id):
    """Allows customers to love/save an item to their wishlist instead of leaving it in cart."""
    product = get_object_or_404(Product, id=product_id, is_active=True)
    wishlist_item, created = Wishlist.objects.get_or_create(customer=request.user, product=product)
    if not created:
        wishlist_item.delete()
    return redirect(request.META.get('HTTP_REFERER', 'store_home'))
import traceback
import random
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from .models import (
    StoreGlobalSetting, StoreOrder, StoreReturnRequest, 
    UserAddressBook, Product, StoreOrderItem, StoreInvoice, 
    CompanyBankAccount, ActivityAuditLog, Notification, StoreRating
)
@login_required
def cart_view(request):
    """Displays cart items alongside strictly sequenced categorized customer orders hub context."""
    try:
        cart = request.session.get('store_cart', {})
        items_subtotal = sum(float(item['price']) * int(item['quantity']) for item in cart.values())
        
        global_settings = StoreGlobalSetting.objects.first()
        
        threshold = float(global_settings.partial_payment_threshold) if global_settings else 0.00
        if items_subtotal >= threshold:
            shipping_fee = 0.00
        else:
            shipping_fee_below = float(global_settings.shipping_fee_below_threshold) if global_settings else 0.00
            shipping_fee = sum(shipping_fee_below * int(item['quantity']) for item in cart.values())
                
        total_amount = items_subtotal + shipping_fee
        requires_full_payment_only = items_subtotal < threshold

        customer_orders = StoreOrder.objects.filter(customer=request.user).order_by('-created_at')
        
        pending_processing_orders = customer_orders.filter(status='pending_processing')
        pending_payment_orders = customer_orders.filter(status__in=['pending_payment', 'partial_payment_submitted'])
        paid_confirmed_orders = customer_orders.filter(status__in=['full_payment_confirmed', 'partial_payment_confirmed', 'balance_payment_submitted'])
        shipping_orders = customer_orders.filter(assigned_rider__isnull=False).exclude(status='delivered')
        delivered_orders = customer_orders.filter(status='delivered')

        # --- CUSTOMER RETURNS & REFUNDS TRACKING ---
        customer_returns = StoreReturnRequest.objects.filter(customer=request.user).order_by('-created_at')
        pending_returns = customer_returns.filter(status='pending')
        approved_returns = customer_returns.filter(status__in=['inspection_approved', 'approved', 'refund_processing'])
        completed_returns = customer_returns.filter(status='refund_completed')
        rejected_returns = customer_returns.filter(status='rejected')

        context = {
            'cart': cart,
            'items_subtotal': items_subtotal,
            'shipping_fee': shipping_fee,
            'total_amount': total_amount,
            'requires_full_payment_only': requires_full_payment_only,
            'global_settings': global_settings,
            'pending_processing_orders': pending_processing_orders,
            'pending_payment_orders': pending_payment_orders,
            'paid_confirmed_orders': paid_confirmed_orders,
            'shipping_orders': shipping_orders,
            'delivered_orders': delivered_orders,
            'customer_returns': customer_returns,
            'pending_returns': pending_returns,
            'approved_returns': approved_returns,
            'completed_returns': completed_returns,
            'rejected_returns': rejected_returns,
        }
        
        return render(request, 'store/cart.html', context)
        
    except Exception as e:
        print("--- CART VIEW ERROR ---")
        traceback.print_exc()
        raise e

@login_required
def checkout_view(request):
    """Processes checkout with validation tied to user's registered address book state and uses active red promotional prices and confidential vendor costs."""
    cart = request.session.get('store_cart', {})
    if not cart:
        return redirect('store_home')
        
    global_settings = StoreGlobalSetting.get_settings()
    threshold = float(global_settings.partial_payment_threshold)
    
    # Retrieve user state from address book
    default_address_obj = UserAddressBook.objects.filter(customer=request.user, is_default=True).first()
    if not default_address_obj:
        default_address_obj = UserAddressBook.objects.filter(customer=request.user).first()
        
    user_state = None
    if default_address_obj:
        default_address = f"{default_address_obj.street_address}, {default_address_obj.lga}, {default_address_obj.state}"
        default_phone = default_address_obj.phone_number
        user_state = default_address_obj.state
    else:
        default_address = getattr(request.user, 'address', '')
        default_phone = getattr(request.user, 'phone_number', '')
        user_state = getattr(request.user, 'state', None)
    
    items_subtotal = sum(float(item['price']) * int(item['quantity']) for item in cart.values())
    
    if request.method == 'POST':
        shipping_address = request.POST.get('shipping_address')
        phone_number = request.POST.get('phone_number')
        payment_choice = request.POST.get('payment_choice', 'full')
        
        if items_subtotal >= threshold:
            shipping_fee = 0.00
        else:
            shipping_fee = float(global_settings.shipping_fee_below_threshold) * sum(int(item['quantity']) for item in cart.values())
            payment_choice = 'full'
                
        total_amount = items_subtotal + shipping_fee
        deposit_amount = total_amount
        balance_amount = 0.00
        
        if payment_choice == 'partial' and items_subtotal >= threshold:
            deposit_pct = float(global_settings.default_deposit_percentage) / 100.0
            deposit_amount = total_amount * deposit_pct
            balance_amount = total_amount - deposit_amount
        else:
            payment_choice = 'full'

        order_status = 'pending_payment'
        
        order = StoreOrder.objects.create(
            customer=request.user,
            status=order_status,
            items_subtotal=items_subtotal,
            shipping_fee=shipping_fee,
            total_amount=total_amount,
            payment_type=payment_choice,
            deposit_amount=deposit_amount,
            balance_amount=balance_amount,
            shipping_address=shipping_address,
            phone_number=phone_number
        )
        
        for prod_id, item in cart.items():
            product = get_object_or_404(Product, id=int(prod_id), is_active=True)

            # Explicitly force the unit price to be the active red price (promo price or discount price if available)
            effective_unit_price = float(product.promo_price) if product.promo_price else (float(product.discount_price) if product.discount_price else float(product.price))

            # Store item with calculated retail subtotal and confidential vendor unit cost
            StoreOrderItem.objects.create(
                order=order,
                product=product,
                quantity=int(item['quantity']),
                unit_price=effective_unit_price,
                total_price=effective_unit_price * int(item['quantity']),
                vendor_unit_price=product.vendor_price  # Pulls main confidential vendor price automatically
            )
            product.stock_quantity -= int(item['quantity'])
            product.save()
            
        # Run order level vendor calculation update method
        if hasattr(order, 'update_vendor_cost'):
            order.update_vendor_cost()
            
        ActivityAuditLog.objects.create(
            user=request.user,
            role='customer',
            action=f"Created Store Order #{order.id} with {payment_choice.upper()} payment, items subtotal ₦{items_subtotal}, shipping ₦{shipping_fee} & deducted warehouse inventory."
        )

        Notification.objects.create(
            user=request.user,
            message=f"Your Store Order #{order.id} has been created. Please complete your payment confirmation."
        )
            
        request.session['store_cart'] = {}
        return redirect('store_order_detail', order_id=order.id)
        
    show_partial_option = items_subtotal >= threshold

    context = {
        'default_address': default_address,
        'default_phone': default_phone,
        'show_partial_option': show_partial_option,
        'items_subtotal': items_subtotal,
        'global_settings': global_settings,
        'saved_addresses': UserAddressBook.objects.filter(customer=request.user),
        'user_state': user_state,
    }
    return render(request, 'store/checkout.html', context)

@login_required
def store_order_detail_view(request, order_id):
    """Customer view to check order status, live pipeline stages, settle remaining balance, and view live map tracker."""
    order = get_object_or_404(StoreOrder, id=order_id, customer=request.user)
    active_bank = CompanyBankAccount.objects.filter(is_active=True).first()
    global_settings = StoreGlobalSetting.get_settings()
    
    if order.payment_type == 'partial':
        deposit_rate = float(global_settings.default_deposit_percentage)
        if hasattr(order, 'global_settings') and order.global_settings:
            deposit_rate = float(order.global_settings.partial_payment_percentage)
            
        order.deposit_amount = (float(order.total_amount) * deposit_rate) / 100
        order.balance_amount = float(order.total_amount) - order.deposit_amount
    else:
        order.deposit_amount = float(order.total_amount)
        order.balance_amount = 0.00

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'mark_paid':
            if order.payment_type == 'partial':
                order.status = 'partial_payment_submitted'
            else:
                order.status = 'pending_payment'
            order.save()
            
            Notification.objects.create(
                user=request.user,
                message=f"Payment notice for Order #{order.id} sent to Finance for verification."
            )
        elif action == 'pay_balance':
            order.status = 'balance_payment_submitted' 
            order.save()
            
            Notification.objects.create(
                user=request.user,
                message=f"Balance payment notice for Order #{order.id} sent to Finance for verification."
            )
        elif action == 'submit_rating':
            score = int(request.POST.get('rating_score', 5))
            comment = request.POST.get('comment', '')
            for item in order.items.all():
                StoreRating.objects.create(
                    product=item.product,
                    customer=request.user,
                    rating_score=score,
                    comment=comment
                )
                
        return redirect('store_order_detail', order_id=order.id)

    if order.status in ['full_payment_confirmed', 'order_ready', 'picked_up', 'arrived_at_customer', 'delivered'] or (order.payment_type == 'partial' and order.balance_paid and order.status != 'balance_payment_submitted'):
        if not order.invoices.exists():
            inv_num = f"S-INV-FULL-{order.id}-{random.randint(1000, 9999)}"
            StoreInvoice.objects.create(
                order=order,
                invoice_number=inv_num,
                invoice_type='full',
                amount_billed=order.total_amount
            )

    invoices = order.invoices.all()
    context = {
        'order': order, 
        'invoices': invoices,
        'active_bank': active_bank,
        'global_settings': global_settings,
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/order_detail.html', context)

    return render(request, 'store/order_detail.html', context)

@login_required
def customer_invoices_view(request, order_id=None):
    """Displays single or multiple customer invoices/orders."""
    global_settings = StoreGlobalSetting.objects.first()
    
    if order_id:
        order = get_object_or_404(StoreOrder, id=order_id, customer=request.user)
        invoice = StoreInvoice.objects.filter(order=order).first()
        orders = [order]
        invoices = [invoice] if invoice else []
    else:
        orders = StoreOrder.objects.filter(customer=request.user).order_by('-created_at')
        invoices = StoreInvoice.objects.filter(order__customer=request.user).order_by('-generated_at')

    return render(request, 'store/customer_invoices.html', {
        'invoice': invoices.first() if hasattr(invoices, 'first') else (invoices[0] if invoices else None),
        'order': orders[0] if orders else None,
        'invoices': invoices, 
        'orders': orders,
        'global_settings': global_settings
    })


def customer_wishlist_view(request):
    wishlist_items = Wishlist.objects.filter(customer=request.user)
    return render(request, 'store/customer_wishlist.html', {'wishlist_items': wishlist_items})


@login_required
def customer_browsing_history_view(request):
    history = BrowsingHistory.objects.filter(customer=request.user).order_by('-viewed_at')
    return render(request, 'store/browsing_history.html', {'history': history})


# ==========================================
# STAFF & EXECUTIVE DASHBOARDS
# ==========================================

from .models import Product, Category, Vendor, PromoTheme, ActivityAuditLog, ProductImage, ProductVideo
@login_required
def store_keeper_dashboard(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['store_keeper', 'general_manager']:
        return redirect('store_home')
        
    # Check if user has management privileges (GM, CEO, or Superuser)
    is_management = request.user.is_superuser or user_role == 'general_manager'

    keeper_state = getattr(request.user, 'state', None)
    if not keeper_state:
        keeper_address = request.user.saved_addresses.filter(is_default=True).first() or request.user.saved_addresses.first()
        if keeper_address:
            keeper_state = keeper_address.state

    products = Product.objects.all().order_by('-id')
    categories = Category.objects.all()
    vendors = Vendor.objects.all().order_by('name')
    promo_themes = PromoTheme.objects.all().order_by('-id')
    
    if request.method == 'POST':
        if 'delete_product' in request.POST:
            prod_id = request.POST.get('product_id')
            product = get_object_or_404(Product, id=prod_id)
            prod_name = product.name
            product.delete()
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'store_keeper', action=f"Deleted product permanently: {prod_name}")
            return redirect('store_keeper_dashboard')

        action = request.POST.get('action') or ('add_product' if request.POST.get('add_product') == '1' else 'edit_product')
        
        # Restrict Category creation to Management/Superuser
        if action == 'add_category':
            if not is_management:
                return redirect('store_keeper_dashboard')
            cat_name = request.POST.get('category_name')
            if cat_name:
                Category.objects.get_or_create(name=cat_name, defaults={'slug': cat_name.lower().replace(' ', '-')})
            return redirect('store_keeper_dashboard')

        # Restrict Vendor creation to Management/Superuser
        elif action == 'add_vendor':
            if not is_management:
                return redirect('store_keeper_dashboard')
            v_name = request.POST.get('vendor_name')
            if v_name:
                Vendor.objects.get_or_create(name=v_name)
            return redirect('store_keeper_dashboard')
            
        elif action == 'add_product':
            name = request.POST.get('name')
            category_id = request.POST.get('category_id')
            vendor_id = request.POST.get('vendor_id')
            description = request.POST.get('description')
            price = request.POST.get('price')
            discount_price = request.POST.get('discount_price') or None
            vendor_price = request.POST.get('vendor_price') or '0.00' if is_management else '0.00'
            promo_price = request.POST.get('promo_price') or None
            promo_theme_id = request.POST.get('promo_theme_id') or None
            stock_quantity = request.POST.get('stock_quantity')
            image = request.FILES.get('image')
            
            # Restrict internal tags and partial payments to management
            internal_brand_tag = request.POST.get('internal_brand_tag', '') if is_management else ''
            allow_partial = True if (is_management and request.POST.get('allow_partial_payment') == 'on') else False
            deposit_pct = int(request.POST.get('partial_deposit_percentage', 80)) if is_management else 80
            
            product = Product.objects.create(
                name=name,
                slug=f"{name.lower().replace(' ', '-')}-{random.randint(100,999)}",
                category_id=int(category_id) if category_id and str(category_id).isdigit() else None,
                vendor_id=int(vendor_id) if vendor_id and str(vendor_id).isdigit() else None,
                description=description,
                price=price,
                discount_price=discount_price,
                vendor_price=vendor_price,
                promo_price=promo_price,
                promo_theme_id=int(promo_theme_id) if promo_theme_id and str(promo_theme_id).isdigit() else None,
                stock_quantity=stock_quantity,
                image=image,
                internal_brand_tag=internal_brand_tag,
                allow_partial_payment=allow_partial,
                partial_deposit_percentage=deposit_pct
            )
            
            for f in request.FILES.getlist('gallery_images') or request.FILES.getlist('additional_images'):
                ProductImage.objects.create(product=product, image=f)
                
            for v in request.FILES.getlist('product_videos') or request.FILES.getlist('demo_videos'):
                ProductVideo.objects.create(product=product, video_file=v)
                
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'store_keeper', action=f"Added new product: {name}")
            return redirect('store_keeper_dashboard')
            
        elif action == 'edit_product':
            prod_id = request.POST.get('product_id')
            product = get_object_or_404(Product, id=prod_id)
            product.name = request.POST.get('name', product.name)
            
            cat_id = request.POST.get('category_id')
            product.category_id = int(cat_id) if cat_id and str(cat_id).isdigit() else product.category_id
            
            # Robustly handle vendor selection updates and clear if unassigned
            new_vendor_id = request.POST.get('vendor_id')
            if new_vendor_id and str(new_vendor_id).isdigit():
                product.vendor_id = int(new_vendor_id)
            else:
                product.vendor = None

            product.description = request.POST.get('description', product.description)
            product.price = request.POST.get('price', product.price)
            product.discount_price = request.POST.get('discount_price') or None
            
            if is_management:
                product.vendor_price = request.POST.get('vendor_price', product.vendor_price)
                product.internal_brand_tag = request.POST.get('internal_brand_tag', product.internal_brand_tag)
                product.allow_partial_payment = True if request.POST.get('allow_partial_payment') == 'on' else False

            product.promo_price = request.POST.get('promo_price') or None
            
            p_theme = request.POST.get('promo_theme_id')
            product.promo_theme_id = int(p_theme) if p_theme and str(p_theme).isdigit() else None
            
            product.stock_quantity = request.POST.get('stock_quantity', product.stock_quantity)
            
            if request.FILES.get('image'):
                product.image = request.FILES.get('image')
                
            product.save()
            
            for f in request.FILES.getlist('gallery_images') or request.FILES.getlist('additional_images'):
                ProductImage.objects.create(product=product, image=f)

            for v in request.FILES.getlist('product_videos') or request.FILES.getlist('demo_videos'):
                ProductVideo.objects.create(product=product, video_file=v)
            
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'store_keeper', action=f"Edited product inventory: {product.name}")
            return redirect('store_keeper_dashboard')
            
    keeper_context = {
        'products': products, 
        'categories': categories, 
        'vendors': vendors,
        'promo_themes': promo_themes,
        'keeper_state': keeper_state,
        'is_management': is_management,
    }

    return render(request, 'store/dashboards/keeper.html', keeper_context)

@login_required
def export_inventory_excel(request):
    """
    Export warehouse inventory as a CSV/Excel spreadsheet.
    Supports filtering by vendor ID via GET parameter (?vendor=ID).
    """
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['store_keeper', 'general_manager', 'finance']:
        return redirect('store_home')

    vendor_id = request.GET.get('vendor')
    products = Product.objects.all().order_by('name')
    
    filename_suffix = "All_Vendors"
    if vendor_id:
        products = products.filter(vendor_id=vendor_id)
        vendor_obj = Vendor.objects.filter(id=vendor_id).first()
        if vendor_obj:
            filename_suffix = vendor_obj.name.replace(" ", "_")

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Warehouse_Inventory_{filename_suffix}.csv"'

    writer = csv.writer(response)
    # Header matching all requested specs (Vendor Name, Vendor Price, Quantity, Image URL, Description, Discount Price, Regular Price, Marketer/Brand Tag, etc.)
    writer.writerow([
        'Product Name', 
        'Vendor Name', 
        'Quantity in Stock', 
        'Vendor Cost Price (₦)', 
        'Regular Retail Price (₦)', 
        'Discount Price (₦)', 
        'Category', 
        'Internal Brand / Marketer Tag', 
        'Image Link', 
        'Description'
    ])

    for p in products:
        v_name = p.vendor.name if p.vendor else "No Vendor Assigned"
        img_url = request.build_absolute_uri(p.image.url) if p.image else "No Image"
        writer.writerow([
            p.name,
            v_name,
            p.stock_quantity,
            p.vendor_price,
            p.price,
            p.discount_price or '',
            p.category.name if p.category else '',
            p.internal_brand_tag or '',
            img_url,
            p.description
        ])

    ActivityAuditLog.objects.create(
        user=request.user, 
        role=user_role or 'store_keeper', 
        action=f"Exported warehouse inventory spreadsheet (Filter: {filename_suffix})"
    )
    return response

from decimal import Decimal, InvalidOperation
import random
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import (
    StoreOrder, 
    StoreReturnRequest, 
    StoreGlobalSetting, 
    TaxRateProposal, 
    StoreInvoice, 
    ActivityAuditLog,
    Notification
)
@login_required
def store_finance_dashboard(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['store_finance', 'Finance', 'general_manager']:
        return redirect('store_home')
        
    pending_orders = StoreOrder.objects.filter(status__in=['pending_payment', 'partial_payment_submitted', 'balance_payment_submitted'])
    
    # Base querysets for filtering
    all_orders = StoreOrder.objects.all().order_by('-id')
    
    # Filtering for Transactions & Invoice Ledger / Vendor Splits
    filter_date = request.GET.get('date')
    filter_status = request.GET.get('status')
    filter_payment_type = request.GET.get('payment_type')
    filter_search = request.GET.get('search')
    filter_amount = request.GET.get('amount')

    if filter_date:
        all_orders = all_orders.filter(created_at__date=filter_date)
    if filter_status:
        all_orders = all_orders.filter(status=filter_status)
    if filter_payment_type:
        all_orders = all_orders.filter(payment_type=filter_payment_type)
    if filter_amount:
        try:
            all_orders = all_orders.filter(total_amount=Decimal(filter_amount))
        except (InvalidOperation, ValueError):
            pass
    if filter_search:
        all_orders = all_orders.filter(
            Q(customer__username__icontains=filter_search) |
            Q(customer__first_name__icontains=filter_search) |
            Q(customer__last_name__icontains=filter_search) |
            Q(id__icontains=filter_search)
        )

    # Fetch returns approved by management that are waiting for finance settlement
    pending_finance_returns = StoreReturnRequest.objects.filter(status='approved_pending_finance').order_by('-created_at')
    
    # Calculate total refunded amount across all completed/closed refunds
    completed_refunds = StoreReturnRequest.objects.filter(status='refunded_and_closed')
    total_refunds_sum = sum(ref.refund_amount for ref in completed_refunds if ref.refund_amount)

    global_settings = StoreGlobalSetting.get_settings()
    pending_tax_proposals = TaxRateProposal.objects.filter(status='pending').order_by('-created_at')

    # Overall calculation sums across all store orders
    total_revenue_sum = sum(order.total_amount for order in StoreOrder.objects.all())
    recognized_revenue_sum = sum(order.recognized_settled_revenue for order in StoreOrder.objects.all() if hasattr(order, 'recognized_settled_revenue') and order.recognized_settled_revenue)
    total_escrow_liability = sum(order.escrow_liability_amount for order in StoreOrder.objects.all() if hasattr(order, 'escrow_liability_amount') and order.escrow_liability_amount)
    total_automated_tax_dues = sum(order.computed_tax_dues for order in StoreOrder.objects.all() if hasattr(order, 'computed_tax_dues') and order.computed_tax_dues)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'propose_tax_rate':
            new_rate_str = request.POST.get('proposed_vat_percentage')
            if new_rate_str:
                try:
                    proposed_rate = Decimal(new_rate_str)
                    TaxRateProposal.objects.create(
                        proposed_percentage=proposed_rate,
                        proposed_by=request.user,
                        status='pending'
                    )
                    ActivityAuditLog.objects.create(
                        user=request.user, 
                        role=user_role or 'store_finance', 
                        action=f"Proposed new VAT tax rate of {proposed_rate}% awaiting CEO/GM approval."
                    )
                except Exception:
                    pass
            return redirect('store_finance_dashboard')

        order_id = request.POST.get('order_id')
        if order_id:
            order = get_object_or_404(StoreOrder, id=order_id)
            
            if action == 'confirm_payment':
                if order.payment_type == 'partial':
                    order.status = 'order_ready'
                    order.save()
                    Notification.objects.create(user=order.customer, message=f"Finance confirmed your initial partial deposit for Order #{order.id}.")
                else:
                    order.status = 'order_ready'
                    order.save()
                    
                    billed_amt = float(order.total_amount)
                    inv_prefix = 'S-INV-FULL'
                    inv_num = f"{inv_prefix}-{order.id}-{random.randint(1000, 9999)}"
                    
                    StoreInvoice.objects.get_or_create(
                        order=order,
                        defaults={'invoice_number': inv_num, 'invoice_type': 'full', 'amount_billed': billed_amt}
                    )
                    Notification.objects.create(user=order.customer, message=f"Finance confirmed full payment & generated invoice for Order #{order.id}.")
                
            elif action == 'confirm_balance':
                order.status = 'order_ready' 
                order.balance_paid = True
                order.save()
                
                billed_amt = float(order.balance_amount) if hasattr(order, 'balance_amount') and order.balance_amount else (float(order.total_amount) * 0.20)
                inv_num = f"S-INV-BAL-{order.id}-{random.randint(1000, 9999)}"
                
                StoreInvoice.objects.get_or_create(
                    order=order,
                    defaults={'invoice_number': inv_num, 'invoice_type': 'partial_balance', 'amount_billed': billed_amt}
                )
                Notification.objects.create(user=order.customer, message=f"Balance payment for Order #{order.id} verified by Finance. Delivery completion unlocked & invoice generated.")
                
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'store_finance', action=f"Processed payment action '{action}' for Order #{order.id}.")
        return redirect('store_finance_dashboard')
        
    context = {
        'pending_orders': pending_orders, 
        'all_orders': all_orders,
        'pending_finance_returns': pending_finance_returns,
        'global_settings': global_settings,
        'pending_tax_proposals': pending_tax_proposals,
        'total_revenue_sum': total_revenue_sum,
        'recognized_revenue_sum': recognized_revenue_sum,
        'total_escrow_liability': total_escrow_liability,
        'total_automated_tax_dues': total_automated_tax_dues,
        'total_refunds_sum': total_refunds_sum,
    }
    
    return render(request, 'store/dashboards/finance.html', context)

from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import StoreReturnRequest, ActivityAuditLog

@login_required
def finance_refund_settlement_view(request, return_id):
    """Finance view to confirm refund, input amount, attach invoice, and close request."""
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role in ['store_finance', 'Finance', 'general_manager']):
        messages.error(request, "Unauthorized access.")
        return redirect('store_home')

    return_obj = get_object_or_404(StoreReturnRequest, id=return_id)

    if request.method == 'POST':
        try:
            raw_amount = request.POST.get('refund_amount', '0.00')
            refund_amount = Decimal(raw_amount)
        except (InvalidOperation, ValueError):
            messages.error(request, "Invalid refund amount format entered.")
            return redirect('store_finance_dashboard')

        refund_invoice = request.FILES.get('refund_invoice')
        
        return_obj.refund_amount = refund_amount
        if refund_invoice:
            return_obj.refund_invoice = refund_invoice
        return_obj.refund_processed_by = request.user
        return_obj.status = 'refunded_and_closed'
        return_obj.save()

        order = return_obj.order
        # Safely reduce the total order amount if the field is editable
        if hasattr(order, 'total_amount') and order.total_amount is not None:
            order.total_amount = max(Decimal('0.00'), order.total_amount - refund_amount)
        order.status = 'refunded'
        order.save()

        ActivityAuditLog.objects.create(
            user=request.user,
            role=user_role or 'store_finance',
            action=f"Processed refund of ₦{refund_amount} for Return Request #{return_obj.id}. Closed ticket."
        )
        messages.success(request, f"Refund settlement of ₦{refund_amount} completed successfully. Customer receipt generated.")
        
        return redirect('store_finance_dashboard')

    return render(request, 'store/finance_refund_settlement.html', {'return_obj': return_obj})

@login_required
def store_manager_dashboard(request):
    """Manager Dashboard with advanced filtering, staff ID card generation, dynamic rider dropdown assignment, and live fleet tracking."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['manager', 'store_manager', 'general_manager', 'rider', 'worker']:
        return redirect('store_home')
        
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_type', '')
    search_query = request.GET.get('q', '')

    orders_qs = StoreOrder.objects.all().select_related('customer', 'assigned_rider').order_by('-id')
    
    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if payment_filter:
        orders_qs = orders_qs.filter(payment_type=payment_filter)
    if search_query:
        orders_qs = orders_qs.filter(
            Q(id__icontains=search_query) |
            Q(customer__username__icontains=search_query) |
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query) |
            Q(phone_number__icontains=search_query) |
            Q(assigned_rider__username__icontains=search_query) |
            Q(assigned_rider__employee_id__icontains=search_query)
        )

    ready_pool = StoreOrder.objects.filter(status='order_ready', assigned_rider__isnull=True).select_related('customer')
    
    live_fleet_orders = StoreOrder.objects.filter(
        status__in=['picked_up', 'arrived_at_customer', 'order_ready']
    ).select_related('customer', 'assigned_rider').order_by('-updated_at')

    audit_logs = ActivityAuditLog.objects.all().order_by('-timestamp')[:100]
    
    all_staff = User.objects.filter(
        Q(role__in=['manager', 'store_manager', 'general_manager', 'rider', 'worker', 'customer_service']) | Q(is_staff=True)
    ).order_by('-id')

    available_riders = User.objects.filter(
        role='rider'
    ).order_by('first_name', 'username')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'generate_staff_id':
            target_user_id = request.POST.get('target_user_id')
            target_user = get_object_or_404(User, id=target_user_id)
            if not getattr(target_user, 'employee_id', None):
                target_user.employee_id = f"TECH-ID-{target_user.id}-{random.randint(1000, 9999)}"
                target_user.save()
                Notification.objects.create(user=target_user, message=f"Your official Staff ID card number has been generated: {target_user.employee_id}")
                ActivityAuditLog.objects.create(user=request.user, role=user_role or 'manager', action=f"Generated Staff ID {target_user.employee_id} for user {target_user.username}")
                messages.success(request, f"Generated official ID {target_user.employee_id} for {target_user.username}.")
            return redirect('store_manager_dashboard')

        order_id = request.POST.get('order_id')
        order = get_object_or_404(StoreOrder, id=order_id)
        
        if action == 'mark_ready':
            order.status = 'order_ready'
            order.save()
            Notification.objects.create(user=order.customer, message=f"Order #{order.id} is ready for dispatch.")
            messages.success(request, f"Order #{order.id} marked as ready for dispatch.")
            
        elif action == 'assign_rider':
            rider_id = request.POST.get('rider_id')
            rider_input = request.POST.get('rider_input', '').strip()
            
            assigned_rider = None
            if rider_id:
                assigned_rider = User.objects.filter(id=rider_id, role='rider').first()
            elif rider_input:
                assigned_rider = User.objects.filter(
                    Q(role='rider') & (Q(username=rider_input) | Q(employee_id=rider_input))
                ).first()
            
            if assigned_rider:
                order.assigned_rider = assigned_rider
                order.status = 'picked_up'
                order.save()
                Notification.objects.create(user=assigned_rider, message=f"Assigned Store Order #{order.id}.")
                messages.success(request, f"Assigned Order #{order.id} to rider {assigned_rider.username}.")
            else:
                messages.error(request, "Selected rider could not be found or does not have the rider role.")
                        
        elif action == 'claim_order':
            if user_role == 'rider' or request.user.is_superuser:
                order.assigned_rider = request.user
                order.status = 'picked_up'
                order.save()
                Notification.objects.create(user=request.user, message=f"You successfully claimed Order #{order.id}.")
                messages.success(request, f"You claimed Order #{order.id}.")
                
        elif action == 'approve_replacement':
            order.status = 'replacement_approved'
            order.save()
            Notification.objects.create(user=order.customer, message=f"Your replacement request for Order #{order.id} has been approved.")
            messages.success(request, f"Approved replacement for Order #{order.id}.")
            
        elif action == 'reject_return':
            order.status = 'return_rejected'
            order.save()
            Notification.objects.create(user=order.customer, message=f"Your return request for Order #{order.id} was rejected.")
            messages.warning(request, f"Rejected return for Order #{order.id}.")
            
        ActivityAuditLog.objects.create(user=request.user, role=user_role or 'manager', action=f"Actioned Order #{order.id}: {action}")
        return redirect('store_manager_dashboard')
        
    manager_context = {
        'orders': orders_qs, 
        'ready_pool': ready_pool,
        'live_fleet_orders': live_fleet_orders,
        'audit_logs': audit_logs,
        'all_staff': all_staff,
        'available_riders': available_riders,
        'current_status_filter': status_filter,
        'current_payment_filter': payment_filter,
        'current_search': search_query
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/manager.html', manager_context)
        
    return render(request, 'store/dashboards/manager.html', manager_context)


@login_required
def price_discount_approvals_view(request):
    """View for managers/CEO to review and approve product price and discount requests."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['general_manager', 'store_manager', 'store_finance']:
        return redirect('store_home')
        
    context = {}
    return render(request, 'store/dashboards/price_discount_approvals.html', context)


@login_required
def remove_from_cart(request, product_id):
    cart = request.session.get('store_cart', {})
    product_id_str = str(product_id)
    if product_id_str in cart:
        del cart[product_id_str]
        request.session['store_cart'] = cart
        request.session.modified = True
    return redirect('store_cart')


def is_ceo(user):
    return user.is_authenticated and (user.is_superuser or getattr(user, 'role', '') == 'ceo')

@login_required
def impersonate_user(request, user_id):
    """Allows superusers or CEOs to impersonate other users for support or debugging."""
    if not request.user.is_superuser and getattr(request.user, 'role', '') != 'ceo':
        return redirect('store_home')
        
    target_user = get_object_or_404(User, id=user_id)
    login(request, target_user)
    messages.success(request, f"You are now impersonating {target_user.username}.")
    return redirect('store_home')

@login_required
@user_passes_test(is_ceo)
def store_ceo_dashboard(request):
    global_settings = StoreGlobalSetting.get_settings()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_global_settings':
            try:
                global_settings.partial_payment_threshold = float(request.POST.get('partial_payment_threshold', 20000))
                global_settings.default_deposit_percentage = int(request.POST.get('default_deposit_percentage', 80))
                global_settings.shipping_fee_below_threshold = float(request.POST.get('shipping_fee_below_threshold', 3000))
                global_settings.save()
                
                ActivityAuditLog.objects.create(
                    user=request.user, 
                    role=getattr(request.user, 'role', 'ceo'), 
                    action="Updated global store financial thresholds."
                )
                messages.success(request, "Global store settings updated successfully.")
            except Exception as e:
                messages.error(request, f"Error updating settings: {e}")
            return redirect('store_ceo_dashboard')

        elif action == 'update_store_invoice_info':
            try:
                global_settings.store_address = request.POST.get('store_address', global_settings.store_address)
                global_settings.store_email = request.POST.get('store_email', global_settings.store_email)
                global_settings.store_phone = request.POST.get('store_phone', global_settings.store_phone)
                global_settings.save()
                
                ActivityAuditLog.objects.create(
                    user=request.user, 
                    role=getattr(request.user, 'role', 'ceo'), 
                    action="Updated official store invoice identity details."
                )
                messages.success(request, "Official invoice store information updated successfully.")
            except Exception as e:
                messages.error(request, f"Error updating invoice info: {e}")
            return redirect('store_ceo_dashboard')

        elif action == 'approve_tax_proposal':
            proposal_id = request.POST.get('proposal_id')
            proposal = get_object_or_404(TaxRateProposal, id=proposal_id)
            proposal.status = 'approved'
            proposal.approved_by = request.user
            proposal.save()

            global_settings.vat_percentage = proposal.proposed_percentage
            global_settings.save()

            ActivityAuditLog.objects.create(
                user=request.user, 
                role=getattr(request.user, 'role', 'ceo'), 
                action=f"Approved Finance VAT rate proposal of {proposal.proposed_percentage}%."
            )
            messages.success(request, f"Tax proposal of {proposal.proposed_percentage}% successfully approved and set live.")
            return redirect('store_ceo_dashboard')

        elif action == 'reject_tax_proposal':
            proposal_id = request.POST.get('proposal_id')
            proposal = get_object_or_404(TaxRateProposal, id=proposal_id)
            proposal.status = 'rejected'
            proposal.approved_by = request.user
            proposal.save()

            ActivityAuditLog.objects.create(
                user=request.user, 
                role=getattr(request.user, 'role', 'ceo'), 
                action=f"Rejected Finance VAT rate proposal of {proposal.proposed_percentage}%."
            )
            messages.warning(request, "Tax rate proposal rejected.")
            return redirect('store_ceo_dashboard')

        elif action == 'assign_user_role_and_id':
            target_user_id = request.POST.get('target_user_id')
            new_role = request.POST.get('role')
            written_id = request.POST.get('written_official_id', '').strip()
            
            target_user = get_object_or_404(User, id=target_user_id)
            
            # Prevent duplicate employee_id integrity errors across different user records
            if written_id:
                # Check if another user already has this employee_id
                existing_user = User.objects.filter(employee_id=written_id).exclude(id=target_user.id).first()
                if existing_user:
                    messages.error(request, f"Error: Employee ID '{written_id}' is already assigned to another user ({existing_user.username}).")
                    return redirect('store_ceo_dashboard')
                target_user.employee_id = written_id
            else:
                # If role is downgraded to customer or ID is blank, clear it or assign None to prevent unique collisions on empty strings
                target_user.employee_id = None if new_role.lower() == 'customer' else target_user.employee_id

            target_user.role = new_role
            
            # Sync staff status flags automatically based on role transition between customer and staff portals
            if new_role.lower() in ['customer', 'client']:
                target_user.is_staff = False
            else:
                target_user.is_staff = True
                
            target_user.save()
            
            ActivityAuditLog.objects.create(
                user=request.user, 
                role=getattr(request.user, 'role', 'ceo'), 
                action=f"Assigned role '{new_role}' and ID '{target_user.employee_id}' to user {target_user.username}."
            )
            messages.success(request, f"Successfully updated user {target_user.username} role and official ID.")
            return redirect('store_ceo_dashboard')

        elif action == 'save_bank_account':
            acc_id = request.POST.get('account_id')
            bank_name = request.POST.get('bank_name')
            account_number = request.POST.get('account_number')
            account_name = request.POST.get('account_name')
            is_active = True if request.POST.get('is_active') == 'on' else False

            if is_active:
                CompanyBankAccount.objects.all().update(is_active=False)

            if acc_id:
                bank_acc = get_object_or_404(CompanyBankAccount, id=acc_id)
                bank_acc.bank_name = bank_name
                bank_acc.account_number = account_number
                bank_acc.account_name = account_name
                bank_acc.is_active = is_active
                bank_acc.save()
            else:
                CompanyBankAccount.objects.create(
                    bank_name=bank_name,
                    account_number=account_number,
                    account_name=account_name,
                    is_active=is_active
                )
            messages.success(request, "Company bank account configuration saved successfully.")
            return redirect('store_ceo_dashboard')

        elif action == 'ceo_override_stage':
            order_id = request.POST.get('order_id')
            new_status = request.POST.get('new_status')
            order = get_object_or_404(StoreOrder, id=order_id)
            order.status = new_status
            order.save()
            
            ActivityAuditLog.objects.create(
                user=request.user, 
                role=getattr(request.user, 'role', 'ceo'), 
                action=f"CEO override: Forced Store Order #{order.id} status to '{new_status}'."
            )
            messages.success(request, f"Store Order #{order.id} stage successfully updated.")
            return redirect('store_ceo_dashboard')

        elif action == 'delete_order':
            order_id = request.POST.get('order_id')
            order = get_object_or_404(StoreOrder, id=order_id)
            order_info = f"#{order.id}"
            order.delete()
            
            ActivityAuditLog.objects.create(
                user=request.user, 
                role=getattr(request.user, 'role', 'ceo'), 
                action=f"CEO deleted job/order {order_info} permanently."
            )
            messages.success(request, f"Job/Order {order_info} has been permanently removed.")
            return redirect('store_ceo_dashboard')

        elif action == 'save_promo_theme':
            promo_theme_id = request.POST.get('promo_theme_id')
            new_theme_name = request.POST.get('new_theme_name', '').strip()
            accent_color = request.POST.get('accent_color', '#ffc107')
            background_color = request.POST.get('background_color', '#111111')
            text_color = request.POST.get('text_color', '#ffffff')

            PromoTheme.objects.all().update(is_active=False)

            if new_theme_name:
                PromoTheme.objects.create(
                    name=new_theme_name,
                    accent_color=accent_color,
                    background_color=background_color,
                    text_color=text_color,
                    is_active=True
                )
            elif promo_theme_id:
                theme = get_object_or_404(PromoTheme, id=promo_theme_id)
                theme.accent_color = accent_color
                theme.background_color = background_color
                theme.text_color = text_color
                theme.is_active = True
                theme.save()

            messages.success(request, "Store celebration theme colors and settings applied.")
            return redirect('store_ceo_dashboard')

        elif action == 'reset_theme':
            PromoTheme.objects.all().update(is_active=False)
            messages.success(request, "Store theme reset to default configuration.")
            return redirect('store_ceo_dashboard')

    all_orders = StoreOrder.objects.all()
    total_revenue_val = sum(o.total_amount for o in all_orders)
    recognized_revenue_sum = sum(o.recognized_settled_revenue for o in all_orders)
    total_escrow_liability = sum(o.escrow_liability_amount for o in all_orders)
    estimated_total_tax = sum(o.computed_tax_dues for o in all_orders)
    pending_tax_proposals = TaxRateProposal.objects.filter(status='pending').order_by('-created_at')

    rider_delivery_logs = StoreOrder.objects.filter(delivery_proof_photo__isnull=False).order_by('-updated_at')

    context = {
        'total_revenue': total_revenue_val,
        'recognized_revenue': recognized_revenue_sum,
        'total_escrow_liability': total_escrow_liability,
        'estimated_tax': estimated_total_tax,
        'pending_tax_proposals': pending_tax_proposals,
        'recent_orders': all_orders.order_by('-created_at')[:15],
        'all_users': User.objects.all().order_by('-date_joined'),
        'bank_accounts': CompanyBankAccount.objects.all(),
        'global_settings': global_settings,
        'promo_themes': PromoTheme.objects.all(),
        'active_theme': PromoTheme.objects.filter(is_active=True).first(),
        'all_ratings': StoreRating.objects.all().order_by('-created_at')[:20],
        'audit_logs': ActivityAuditLog.objects.all().order_by('-timestamp')[:50],
        'rider_delivery_logs': rider_delivery_logs,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/ceo.html', context)
        
    return render(request, 'store/dashboards/ceo.html', context)


@login_required
def store_rider_dashboard(request):
    """Rider dashboard with GPS telemetry coordinate broadcasting and ready pool order claiming."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['rider', 'general_manager']:
        return redirect('store_home')
        
    assigned_deliveries = StoreOrder.objects.filter(assigned_rider=request.user).order_by('-id')
    ready_pool_orders = StoreOrder.objects.filter(status='order_ready', assigned_rider__isnull=True).order_by('-id')
    rider_history = assigned_deliveries.exclude(delivery_proof_photo='')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        order_id = request.POST.get('order_id')
        
        if action == 'claim_order':
            order = get_object_or_404(StoreOrder, id=order_id, status='order_ready', assigned_rider__isnull=True)
            order.assigned_rider = request.user
            order.status = 'picked_up'
            order.save()
            Notification.objects.create(user=request.user, message=f"You successfully claimed Order #{order.id} from the pool.")
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'rider', action=f"Claimed Order #{order.id} from ready pool")
            return redirect('store_rider_dashboard')

        order = get_object_or_404(StoreOrder, id=order_id, assigned_rider=request.user)
        
        if action == 'pick_goods':
            order.status = 'picked_up'
            order.is_location_live = True
            order.rider_latitude = request.POST.get('latitude', 6.5244)
            order.rider_longitude = request.POST.get('longitude', 3.3792)
            order.save()
            Notification.objects.create(user=order.customer, message=f"Rider picked up Order #{order.id}. Tracking live.")
        elif action == 'update_location':
            order.rider_latitude = request.POST.get('latitude')
            order.rider_longitude = request.POST.get('longitude')
            order.is_location_live = True
            order.save()
            return JsonResponse({'status': 'success', 'lat': order.rider_latitude, 'lng': order.rider_longitude})
        elif action == 'mark_arrived':
            order.status = 'arrived_at_customer'
            order.is_location_live = False
            order.save()
            Notification.objects.create(user=order.customer, message=f"Rider arrived at destination for Order #{order.id}.")
        elif action == 'mark_delivered':
            if order.payment_type == 'partial' and not order.balance_paid:
                Notification.objects.create(user=request.user, message=f"Cannot complete Order #{order.id}: Remaining balance has not been verified by Finance yet.")
            else:
                if request.FILES.get('delivery_proof_photo'):
                    order.delivery_proof_photo = request.FILES.get('delivery_proof_photo')
                order.status = 'delivered'
                order.is_location_live = False
                order.save()
                Notification.objects.create(user=order.customer, message=f"Order #{order.id} delivered successfully.")

        ActivityAuditLog.objects.create(user=request.user, role=user_role or 'rider', action=f"Updated status for Order #{order.id}")
        return redirect('store_rider_dashboard')
        
    rider_context = {
        'assigned_deliveries': assigned_deliveries,
        'ready_pool_orders': ready_pool_orders,
        'rider_history': rider_history
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/rider.html', rider_context)
        
    return render(request, 'store/dashboards/rider.html', rider_context)


@login_required
def ceo_gm_dispatch_logs_view(request):
    """Exclusive Dispatch & Process Logs view for CEO and General Manager with filters and exports."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['ceo', 'general_manager']:
        return redirect('store_home')

    orders = StoreOrder.objects.all().order_by('-created_at')
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status=status_filter)

    context = {
        'orders': orders,
        'status_filter': status_filter
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/dispatch_logs.html', context)

    return render(request, 'store/dashboards/dispatch_logs.html', context)


@login_required
def export_audit_logs_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['manager', 'store_manager', 'store_finance', 'general_manager']:
        return redirect('store_home')

    selected_year = int(request.GET.get('year', timezone.now().year))

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    orders = StoreOrder.objects.filter(created_at__year=selected_year).order_by('created_at')

    months_data = {m: [] for m in range(1, 13)}
    for order in orders:
        months_data[order.created_at.month].append(order)

    has_data_written = False

    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num]
        month_orders = months_data[month_num]

        ws = wb.create_sheet(title=f"{month_name} {selected_year}")
        has_data_written = True

        ws.append([f"TechsNi Store Financial & Audit Ledger - {month_name} {selected_year}"])
        ws.append([]) 

        headers = [
            'Order ID', 'Customer', 'Marketer Ref', 'Payment Type', 
            'Date & Time', 'Total Amount (₦)', 'VAT Dues (₦)', 
            'Vendor Cost (₦)', 'Net Store Revenue (₦)', 'Marketer Commission (₦)', 'Status'
        ]
        ws.append(headers)

        month_total_revenue = Decimal('0.00')
        month_total_vat = Decimal('0.00')
        month_total_vendor_cost = Decimal('0.00')
        month_total_net_revenue = Decimal('0.00')
        month_total_commission = Decimal('0.00')

        for order in month_orders:
            cust_name = order.customer.username if order.customer else 'Unknown'
            marketer_ref = order.customer.referred_by.username if hasattr(order.customer, 'referred_by') and order.customer.referred_by else 'None'
            
            vat_val = order.computed_tax_dues
            vendor_cost = getattr(order, 'vendor_cost_amount', Decimal('0.00')) or Decimal('0.00')
            net_rev = order.total_amount - vendor_cost
            
            marketer_comm = Decimal('0.00')
            if hasattr(order.customer, 'referred_by') and order.customer.referred_by:
                commission_pct = getattr(order.customer.referred_by, 'commission_percentage', Decimal('5.00'))
                marketer_comm = (net_rev * (Decimal(str(commission_pct)) / Decimal('100.00'))).quantize(Decimal('0.01'))

            month_total_revenue += order.total_amount
            month_total_vat += vat_val
            month_total_vendor_cost += vendor_cost
            month_total_net_revenue += net_rev
            month_total_commission += marketer_comm

            row_data = [
                f"#{order.id}",
                cust_name,
                marketer_ref,
                order.get_payment_type_display(),
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                float(order.total_amount),
                float(vat_val),
                float(vendor_cost),
                float(net_rev),
                float(marketer_comm),
                order.get_status_display()
            ]
            ws.append(row_data)

        ws.append([])
        ws.append([
            "MONTHLY TOTALS", "", "", "", "", 
            float(month_total_revenue), 
            float(month_total_vat), 
            float(month_total_vendor_cost), 
            float(month_total_net_revenue), 
            float(month_total_commission), 
            ""
        ])

    if not has_data_written:
        fallback_ws = wb.create_sheet(title="Overview")
        fallback_ws.append(["No transaction records found for year", selected_year])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TechsNi_Audit_Ledger_{selected_year}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_ceo_transactions_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role != 'general_manager':
        return redirect('store_home')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="TechsNi_Store_Transactions.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Items Subtotal', 'Shipping Fee', 'Total Amount', 'Payment Type', 'Status', 'Date'])
    for o in StoreOrder.objects.all().order_by('-created_at'):
        writer.writerow([o.id, o.customer.username, o.items_subtotal, o.shipping_fee, o.total_amount, o.payment_type, o.status, o.created_at])
    return response


@login_required
def export_ceo_inventory_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role != 'general_manager':
        return redirect('store_home')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="TechsNi_Warehouse_Inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(['Product ID', 'Product Name', 'Category', 'Price', 'Stock', 'Status'])
    for p in Product.objects.all().order_by('id'):
        writer.writerow([p.id, p.name, p.category.name if p.category else 'Uncategorized', p.price, p.stock_quantity, p.is_active])
    return response


@login_required
def user_profile_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_address':
            state = request.POST.get('state')
            lga = request.POST.get('lga')
            street_address = request.POST.get('street_address')
            phone_number = request.POST.get('phone_number')
            is_default = request.POST.get('is_default') == 'on'
            
            if is_default:
                UserAddressBook.objects.filter(customer=request.user).update(is_default=False)
                
            UserAddressBook.objects.create(
                customer=request.user,
                state=state,
                lga=lga,
                street_address=street_address,
                phone_number=phone_number,
                is_default=is_default or not UserAddressBook.objects.filter(customer=request.user).exists()
            )
            return redirect('user_profile')
            
        else:
            request.user.first_name = request.POST.get('first_name', request.user.first_name)
            request.user.last_name = request.POST.get('last_name', request.user.last_name)
            request.user.email = request.POST.get('email', request.user.email)
            request.user.save()
            return redirect('user_profile')
            
    saved_addresses = UserAddressBook.objects.filter(customer=request.user).order_by('-is_default', '-id')
    return render(request, 'store/user_profile.html', {
        'policies': CompanyPolicy.objects.all(),
        'saved_addresses': saved_addresses,
    })


@login_required
def delete_account_view(request):
    if request.method == 'POST':
        request.user.delete()
        return redirect('login')
    return render(request, 'store/delete_account_confirmation.html')


@login_required
def portal_gateway_view(request):
    return render(request, 'store/portal_gateway.html')

@login_required
def add_to_cart_view(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    
    if product.stock_quantity <= 0:
        return redirect(request.META.get('HTTP_REFERER', 'store_home'))

    cart = request.session.get('store_cart', {})
    try:
        requested_qty = int(request.POST.get('quantity', 1))
    except ValueError:
        requested_qty = 1

    str_id = str(product_id)
    current_qty = cart.get(str_id, {}).get('quantity', 0)

    if product.stock_quantity < (current_qty + requested_qty):
        return redirect(request.META.get('HTTP_REFERER', 'store_home'))

    # Explicitly prioritize promo_price, then discount_price, then regular price (Red price logic)
    if product.promo_price and float(product.promo_price) > 0:
        active_price = float(product.promo_price)
    elif product.discount_price and float(product.discount_price) > 0:
        active_price = float(product.discount_price)
    else:
        active_price = float(product.price)

    if str_id in cart:
        cart[str_id]['quantity'] += requested_qty
        cart[str_id]['price'] = active_price
    else:
        cart[str_id] = {
            'name': product.name,
            'price': active_price,
            'quantity': requested_qty,
            'image': product.image.url if product.image else '',
            'allow_partial': product.allow_partial_payment,
            'deposit_percentage': product.partial_deposit_percentage
        }
        
    request.session['store_cart'] = cart
    return redirect('store_cart')
    


@login_required
def update_cart_quantity_view(request, product_id):
    """Allows customers to reduce, increase, or remove individual cart item quantities dynamically."""
    cart = request.session.get('store_cart', {})
    str_id = str(product_id)

    if str_id in cart:
        action = request.POST.get('action')
        
        if action == 'decrease':
            cart[str_id]['quantity'] -= 1
            if cart[str_id]['quantity'] <= 0:
                del cart[str_id]
        elif action == 'increase':
            product = get_object_or_404(Product, id=product_id, is_active=True)
            if cart[str_id]['quantity'] < product.stock_quantity:
                cart[str_id]['quantity'] += 1
        elif action == 'remove':
            del cart[str_id]
            
        request.session['store_cart'] = cart

    return redirect('store_cart')


# ==========================================
# NOTIFICATIONS & ADDITIONAL CUSTOMER VIEWS
# ==========================================

@login_required
def notification_list_view(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    notifications.filter(is_read=False).update(is_read=True)
    return render(request, 'store/notifications.html', {
        'notifications': notifications,
    })


@login_required
def clear_notifications_view(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user).delete()
    return redirect('notification_list')


@login_required
def customer_orders_list_view(request):
    status_filter = request.GET.get('status')
    orders = StoreOrder.objects.filter(customer=request.user).order_by('-created_at')
    
    if status_filter == 'pending':
        orders = orders.filter(status__in=['pending_processing', 'order_ready'])
    elif status_filter == 'pending_payment':
        orders = orders.filter(status__in=['pending_payment', 'partial_payment_submitted'])
    elif status_filter == 'confirmed':
        orders = orders.filter(status__in=['full_payment_confirmed', 'partial_payment_confirmed', 'balance_paid'])
    elif status_filter == 'shipping':
        orders = orders.filter(status__in=['picked_up', 'arrived_at_customer'])

    return render(request, 'store/customer_orders.html', {
        'orders': orders,
        'current_filter': status_filter,
    })


# ==========================================
# COMPANY POLICIES & STATIC INFORMATION
# ==========================================

def company_policy_view(request, policy_slug=None):
    policies = CompanyPolicy.objects.all()
    active_policy = None
    if policy_slug:
        active_policy = get_object_or_404(CompanyPolicy, slug=policy_slug)
    elif policies.exists():
        active_policy = policies.first()

    return render(request, 'store/policies.html', {
        'policies': policies,
        'active_policy': active_policy,
    })


# ==========================================
# SYSTEM UTILITY & ERROR HANDLING HOOKS
# ==========================================

def handler_404_view(request, exception):
    return render(request, 'store/errors/404.html', status=404)


def handler_500_view(request):
    return render(request, 'store/errors/500.html', status=500)


@login_required
def api_sync_marketer(request):
    return JsonResponse({'status': 'success'})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    messages.success(request, f'{product.name} has been added to your cart.')
    return redirect(request.META.get('HTTP_REFERER', 'store:product_list'))


@login_required
def toggle_wishlist(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    wishlist_item, created = Wishlist.objects.get_or_create(
        customer=request.user, product=product
    )
    if not created:
        wishlist_item.delete()
        messages.info(request, f'{product.name} was removed from your wishlist.')
    else:
        messages.success(request, f'{product.name} was added to your wishlist.')
    return redirect(request.META.get('HTTP_REFERER', 'store:product_list'))


@login_required
def submit_product_rating_view(request, product_id):
    """Handles customer star ratings and review submissions for products."""
    product = get_object_or_404(Product, id=product_id, is_active=True)
    
    if request.method == 'POST':
        try:
            rating_value = int(request.POST.get('rating', 5))
        except ValueError:
            rating_value = 5
            
        comment_text = request.POST.get('comment', '').strip()
        
        StoreRating.objects.update_or_create(
            user=request.user,
            product=product,
            defaults={
                'rating': max(1, min(5, rating_value)),
                'comment': comment_text
            }
        )
        
    return redirect(request.META.get('HTTP_REFERER', 'store_home'))


@login_required
def general_manager_dashboard(request):
    """General Manager Dashboard with high-level operational supervision and customer ratings oversight."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role != 'general_manager':
        return redirect('store_home')
        
    global_settings = StoreGlobalSetting.get_settings()

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve_tax_proposal':
            proposal_id = request.POST.get('proposal_id')
            proposal = get_object_or_404(TaxRateProposal, id=proposal_id)
            
            proposal.status = 'approved'
            proposal.approved_by = request.user
            proposal.save()

            global_settings.vat_percentage = proposal.proposed_percentage
            global_settings.save()

            ActivityAuditLog.objects.create(
                user=request.user, 
                role=user_role or 'general_manager', 
                action=f"Approved Finance VAT rate proposal of {proposal.proposed_percentage}%."
            )
            messages.success(request, f"Tax proposal of {proposal.proposed_percentage}% successfully approved and set live.")
            return redirect('general_manager_dashboard')

        elif action == 'reject_tax_proposal':
            proposal_id = request.POST.get('proposal_id')
            proposal = get_object_or_404(TaxRateProposal, id=proposal_id)
            proposal.status = 'rejected'
            proposal.approved_by = request.user
            proposal.save()

            ActivityAuditLog.objects.create(
                user=request.user, 
                role=user_role or 'general_manager', 
                action=f"Rejected Finance VAT rate proposal of {proposal.proposed_percentage}%."
            )
            messages.warning(request, "Tax rate proposal rejected.")
            return redirect('general_manager_dashboard')

    recent_orders = StoreOrder.objects.all().order_by('-created_at')[:100]
    all_users = User.objects.all().order_by('-id')
    audit_logs = ActivityAuditLog.objects.all().order_by('-timestamp')[:200]
    store_ratings = StoreRating.objects.all().order_by('-created_at')[:50]
    pending_tax_proposals = TaxRateProposal.objects.filter(status='pending').order_by('-created_at')

    rider_delivery_logs = StoreOrder.objects.filter(delivery_proof_photo__isnull=False).order_by('-updated_at')

    all_orders_queryset = StoreOrder.objects.all()
    total_revenue = sum(order.total_amount for order in all_orders_queryset)
    recognized_revenue_sum = sum(order.recognized_settled_revenue for order in all_orders_queryset)
    total_escrow_liability = sum(order.escrow_liability_amount for order in all_orders_queryset)
    estimated_tax = sum(order.computed_tax_dues for order in all_orders_queryset)

    gm_context = {
        'recent_orders': recent_orders,
        'all_users': all_users,
        'audit_logs': audit_logs,
        'store_ratings': store_ratings,
        'total_revenue': total_revenue,
        'recognized_revenue': recognized_revenue_sum,
        'total_escrow_liability': total_escrow_liability,
        'estimated_tax': estimated_tax,
        'pending_tax_proposals': pending_tax_proposals,
        'global_settings': global_settings,
        'rider_delivery_logs': rider_delivery_logs,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/general_manager.html', gm_context)
        
    return render(request, 'store/dashboards/general_manager.html', gm_context)

from decimal import Decimal
import calendar
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from .models import StoreOrder, Notification, ActivityAuditLog, Product, UserAddressBook, CompanyPolicy, Wishlist, StoreRating, StoreInvoice, CompanyBankAccount, StoreGlobalSetting

@login_required
def store_rider_dashboard(request):
    """Rider dashboard with GPS telemetry coordinate broadcasting, ready pool order claiming, and assigned return tracking."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['rider', 'general_manager']:
        return redirect('store_home')
        
    assigned_deliveries = StoreOrder.objects.filter(assigned_rider=request.user).order_by('-id')
    ready_pool_orders = StoreOrder.objects.filter(status='order_ready', assigned_rider__isnull=True).order_by('-id')
    
    # Fetch return requests assigned to this specific rider
    assigned_returns = StoreReturnRequest.objects.filter(assigned_rider=request.user).order_by('-created_at')
    
    # Rider upload/view history of uploaded delivery proof files
    rider_history = assigned_deliveries.exclude(delivery_proof_photo='')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        order_id = request.POST.get('order_id')
        
        if action == 'claim_order':
            order = get_object_or_404(StoreOrder, id=order_id, status='order_ready', assigned_rider__isnull=True)
            order.assigned_rider = request.user
            order.status = 'picked_up'
            order.save()
            Notification.objects.create(user=request.user, message=f"You successfully claimed Order #{order.id} from the pool.")
            ActivityAuditLog.objects.create(user=request.user, role=user_role or 'rider', action=f"Claimed Order #{order.id} from ready pool")
            return redirect('store_rider_dashboard')

        order = get_object_or_404(StoreOrder, id=order_id, assigned_rider=request.user)
        
        if action == 'pick_goods':
            order.status = 'picked_up'
            order.is_location_live = True
            order.rider_latitude = request.POST.get('latitude', 6.5244)
            order.rider_longitude = request.POST.get('longitude', 3.3792)
            order.save()
            Notification.objects.create(user=order.customer, message=f"Rider picked up Order #{order.id}. Tracking live.")
        elif action == 'update_location':
            order.rider_latitude = request.POST.get('latitude')
            order.rider_longitude = request.POST.get('longitude')
            order.is_location_live = True
            order.save()
            return JsonResponse({'status': 'success', 'lat': order.rider_latitude, 'lng': order.rider_longitude})
        elif action == 'mark_arrived':
            order.status = 'arrived_at_customer'
            order.is_location_live = False
            order.save()
            Notification.objects.create(user=order.customer, message=f"Rider arrived at destination for Order #{order.id}.")
        elif action == 'mark_delivered':
            # Block delivery if partial payment balance is unpaid/unconfirmed by finance
            if order.payment_type == 'partial' and not order.balance_paid:
                Notification.objects.create(user=request.user, message=f"Cannot complete Order #{order.id}: Remaining balance has not been verified by Finance yet.")
            else:
                if request.FILES.get('delivery_proof_photo'):
                    order.delivery_proof_photo = request.FILES.get('delivery_proof_photo')
                order.status = 'delivered'
                order.is_location_live = False
                order.save()
                Notification.objects.create(user=order.customer, message=f"Order #{order.id} delivered successfully.")

        ActivityAuditLog.objects.create(user=request.user, role=user_role or 'rider', action=f"Updated status for Order #{order.id}")
        return redirect('store_rider_dashboard')
        
    context = {
        'assigned_deliveries': assigned_deliveries,
        'ready_pool_orders': ready_pool_orders,
        'assigned_returns': assigned_returns, # <-- Added assigned returns to context
        'rider_history': rider_history
    }
    
    # Automatically detect and return partial template or full render depending on AJAX ping request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/rider.html', context)
        
    return render(request, 'store/dashboards/rider.html', context)


@login_required
def ceo_gm_dispatch_logs_view(request):
    """Exclusive Dispatch & Process Logs view for CEO and General Manager with filters, exports, and automatic live-refresh."""
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['ceo', 'general_manager']:
        return redirect('store_home')

    orders = StoreOrder.objects.all().order_by('-created_at')

    # Filtering options
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status=status_filter)

    # Export handlers (PDF / Excel placeholders or query triggers)
    export_format = request.GET.get('export')
    if export_format == 'excel':
        pass
    elif export_format == 'pdf':
        pass

    context = {
        'orders': orders,
        'status_filter': status_filter
    }

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'store/dashboards/dispatch_logs.html', context)

    return render(request, 'store/dashboards/dispatch_logs.html', context)


# ==========================================
# EXPORT ENGINES & PROFILES
# ==========================================

@login_required
def export_audit_logs_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role not in ['manager', 'store_manager', 'store_finance', 'general_manager']:
        return redirect('store_home')

    # Optional year filter from query params (defaults to current year)
    selected_year = int(request.GET.get('year', timezone.now().year))

    wb = Workbook()
    # Remove default sheet so we can build month-by-month tabs dynamically
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Fetch orders for the selected year
    orders = StoreOrder.objects.filter(created_at__year=selected_year).order_by('created_at')

    # Group orders by month (1 to 12)
    months_data = {m: [] for m in range(1, 13)}
    for order in orders:
        months_data[order.created_at.month].append(order)

    has_data_written = False

    for month_num in range(1, 13):
        month_name = calendar.month_name[month_num]
        month_orders = months_data[month_num]

        # Create a sheet for each month
        ws = wb.create_sheet(title=f"{month_name} {selected_year}")
        has_data_written = True

        # Header Information Block
        ws.append([f"TechsNi Store Financial & Audit Ledger - {month_name} {selected_year}"])
        ws.append([])  # Blank row

        # Table Column Headers
        headers = [
            'Order ID', 'Customer', 'Marketer Ref', 'Payment Type', 
            'Date & Time', 'Total Amount (₦)', 'VAT Dues (₦)', 
            'Vendor Cost (₦)', 'Net Store Revenue (₦)', 'Marketer Commission (₦)', 'Status'
        ]
        ws.append(headers)

        month_total_revenue = Decimal('0.00')
        month_total_vat = Decimal('0.00')
        month_total_vendor_cost = Decimal('0.00')
        month_total_net_revenue = Decimal('0.00')
        month_total_commission = Decimal('0.00')

        for order in month_orders:
            cust_name = order.customer.username if order.customer else 'Unknown'
            marketer_ref = order.customer.referred_by.username if hasattr(order.customer, 'referred_by') and order.customer.referred_by else 'None'
            
            # Calculations
            vat_val = order.computed_tax_dues
            vendor_cost = getattr(order, 'vendor_cost_amount', Decimal('0.00')) or Decimal('0.00')
            net_rev = order.total_amount - vendor_cost
            
            # Marketer Commission computation (assuming 5% or custom tied percentage if available)
            marketer_comm = Decimal('0.00')
            if hasattr(order.customer, 'referred_by') and order.customer.referred_by:
                commission_pct = getattr(order.customer.referred_by, 'commission_percentage', Decimal('5.00'))
                marketer_comm = (net_rev * (Decimal(str(commission_pct)) / Decimal('100.00'))).quantize(Decimal('0.01'))

            month_total_revenue += order.total_amount
            month_total_vat += vat_val
            month_total_vendor_cost += vendor_cost
            month_total_net_revenue += net_rev
            month_total_commission += marketer_comm

            row_data = [
                f"#{order.id}",
                cust_name,
                marketer_ref,
                order.get_payment_type_display(),
                order.created_at.strftime("%Y-%m-%d %H:%M"),
                float(order.total_amount),
                float(vat_val),
                float(vendor_cost),
                float(net_rev),
                float(marketer_comm),
                order.get_status_display()
            ]
            ws.append(row_data)

        # Summary Row at the bottom of each month sheet
        ws.append([])
        ws.append([
            "MONTHLY TOTALS", "", "", "", "", 
            float(month_total_revenue), 
            float(month_total_vat), 
            float(month_total_vendor_cost), 
            float(month_total_net_revenue), 
            float(month_total_commission), 
            ""
        ])

   # Fallback if no orders exist at all for the year
    if not has_data_written:
        fallback_ws = wb.create_sheet(title="Overview")
        fallback_ws.append(["No transaction records found for year", selected_year])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="TechsNi_Audit_Ledger_{selected_year}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_ceo_transactions_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role != 'general_manager':
        return redirect('store_home')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="TechsNi_Store_Transactions.csv"'
    writer = csv.writer(response)
    writer.writerow(['Order ID', 'Customer', 'Items Subtotal', 'Shipping Fee', 'Total Amount', 'Payment Type', 'Status', 'Date'])
    for o in StoreOrder.objects.all().order_by('-created_at'):
        writer.writerow([o.id, o.customer.username, o.items_subtotal, o.shipping_fee, o.total_amount, o.payment_type, o.status, o.created_at])
    return response


@login_required
def export_ceo_inventory_csv(request):
    user_role = getattr(request.user, 'role', '')
    if not request.user.is_superuser and user_role != 'general_manager':
        return redirect('store_home')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="TechsNi_Warehouse_Inventory.csv"'
    writer = csv.writer(response)
    writer.writerow(['Product ID', 'Product Name', 'Category', 'Price', 'Stock', 'Status'])
    for p in Product.objects.all().order_by('id'):
        writer.writerow([p.id, p.name, p.category.name if p.category else 'Uncategorized', p.price, p.stock_quantity, p.is_active])
    return response


@login_required
def user_profile_view(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'add_address':
            state = request.POST.get('state')
            lga = request.POST.get('lga')
            street_address = request.POST.get('street_address')
            phone_number = request.POST.get('phone_number')
            is_default = request.POST.get('is_default') == 'on'
            
            if is_default:
                UserAddressBook.objects.filter(customer=request.user).update(is_default=False)
                
            UserAddressBook.objects.create(
                customer=request.user,
                state=state,
                lga=lga,
                street_address=street_address,
                phone_number=phone_number,
                is_default=is_default or not UserAddressBook.objects.filter(customer=request.user).exists()
            )
            return redirect('user_profile')
            
        else:
            request.user.first_name = request.POST.get('first_name', request.user.first_name)
            request.user.last_name = request.POST.get('last_name', request.user.last_name)
            request.user.email = request.POST.get('email', request.user.email)
            request.user.save()
            return redirect('user_profile')
            
    saved_addresses = UserAddressBook.objects.filter(customer=request.user).order_by('-is_default', '-id')
    return render(request, 'store/user_profile.html', {
        'policies': CompanyPolicy.objects.all(),
        'saved_addresses': saved_addresses,
    })


@login_required
def delete_account_view(request):
    if request.method == 'POST':
        request.user.delete()
        return redirect('login')
    return render(request, 'store/delete_account_confirmation.html')


@login_required
def portal_gateway_view(request):
    return render(request, 'store/portal_gateway.html')


@login_required
def add_to_cart_view(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    
    if product.stock_quantity <= 0:
        return redirect(request.META.get('HTTP_REFERER', 'store_home'))

    cart = request.session.get('store_cart', {})
    try:
        requested_qty = int(request.POST.get('quantity', 1))
    except ValueError:
        requested_qty = 1

    str_id = str(product_id)
    current_qty = cart.get(str_id, {}).get('quantity', 0)

    if product.stock_quantity < (current_qty + requested_qty):
        return redirect(request.META.get('HTTP_REFERER', 'store_home'))

    active_price = float(product.current_active_price)

    if str_id in cart:
        cart[str_id]['quantity'] += requested_qty
        cart[str_id]['price'] = active_price
    else:
        cart[str_id] = {
            'name': product.name,
            'price': active_price,
            'quantity': requested_qty,
            'image': product.image.url if product.image else '',
            'allow_partial': product.allow_partial_payment,
            'deposit_percentage': product.partial_deposit_percentage
        }
        
    request.session['store_cart'] = cart
    return redirect('store_cart')


@login_required
def update_cart_quantity_view(request, product_id):
    """Allows customers to reduce, increase, or remove individual cart item quantities dynamically."""
    cart = request.session.get('store_cart', {})
    str_id = str(product_id)

    if str_id in cart:
        action = request.POST.get('action')
        
        if action == 'decrease':
            cart[str_id]['quantity'] -= 1
            if cart[str_id]['quantity'] <= 0:
                del cart[str_id]
        elif action == 'increase':
            product = get_object_or_404(Product, id=product_id, is_active=True)
            if cart[str_id]['quantity'] < product.stock_quantity:
                cart[str_id]['quantity'] += 1
        elif action == 'remove':
            del cart[str_id]
            
        request.session['store_cart'] = cart

    return redirect('store_cart')


# ==========================================
# NOTIFICATIONS & ADDITIONAL CUSTOMER VIEWS
# ==========================================

@login_required
def notification_list_view(request):
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    notifications.filter(is_read=False).update(is_read=True)
    return render(request, 'store/notifications.html', {
        'notifications': notifications,
    })


@login_required
def clear_notifications_view(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user).delete()
    return redirect('notification_list')


@login_required
def customer_orders_list_view(request):
    status_filter = request.GET.get('status')
    orders = StoreOrder.objects.filter(customer=request.user).order_by('-created_at')
    
    if status_filter == 'pending':
        orders = orders.filter(status__in=['pending_processing', 'order_ready'])
    elif status_filter == 'pending_payment':
        orders = orders.filter(status__in=['pending_payment', 'partial_payment_submitted'])
    elif status_filter == 'confirmed':
        orders = orders.filter(status__in=['full_payment_confirmed', 'partial_payment_confirmed', 'balance_paid'])
    elif status_filter == 'shipping':
        orders = orders.filter(status__in=['picked_up', 'arrived_at_customer'])

    return render(request, 'store/customer_orders.html', {
        'orders': orders,
        'current_filter': status_filter,
    })


# ==========================================
# COMPANY POLICIES & STATIC INFORMATION
# ==========================================

def company_policy_view(request, policy_slug=None):
    policies = CompanyPolicy.objects.all()
    active_policy = None
    if policy_slug:
        active_policy = get_object_or_404(CompanyPolicy, slug=policy_slug)
    elif policies.exists():
        active_policy = policies.first()

    return render(request, 'store/policies.html', {
        'policies': policies,
        'active_policy': active_policy,
    })


# ==========================================
# SYSTEM UTILITY & ERROR HANDLING HOOKS
# ==========================================

def handler_404_view(request, exception):
    return render(request, 'store/errors/404.html', status=404)


def handler_500_view(request):
    return render(request, 'store/errors/500.html', status=500)


@login_required
def api_sync_marketer(request):
    return JsonResponse({'status': 'success'})


@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, id=product_id, is_active=True)
    messages.success(request, f'{product.name} has been added to your cart.')
    return redirect(request.META.get('HTTP_REFERER', 'store:product_list'))


@login_required
def toggle_wishlist(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    wishlist_item, created = Wishlist.objects.get_or_create(
        customer=request.user, product=product
    )
    if not created:
        wishlist_item.delete()
        messages.info(request, f'{product.name} was removed from your wishlist.')
    else:
        messages.success(request, f'{product.name} was added to your wishlist.')
    return redirect(request.META.get('HTTP_REFERER', 'store:product_list'))


@login_required
def submit_product_rating_view(request, product_id):
    """Handles customer star ratings and review submissions for products."""
    product = get_object_or_404(Product, id=product_id, is_active=True)
    
    if request.method == 'POST':
        try:
            rating_value = int(request.POST.get('rating', 5))
        except ValueError:
            rating_value = 5
            
        comment_text = request.POST.get('comment', '').strip()
        
        # Check if user already reviewed this product, update it or create new
        StoreRating.objects.update_or_create(
            user=request.user,
            product=product,
            defaults={
                'rating': max(1, min(5, rating_value)), # Ensure rating is between 1 and 5
                'comment': comment_text
            }
        )
        
    return redirect(request.META.get('HTTP_REFERER', 'store_home'))




from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from django.http import HttpResponse
from django.db.models import Q
import csv

from django.contrib.auth import get_user_model
User = get_user_model()

from .models import (
    StoreOrder, StoreReturnRequest, StoreReturnPolicy, 
    StoreReturnItem, RefundReason, ActivityAuditLog
)
from .forms import CustomerReturnRequestForm

@login_required
def request_store_refund_view(request, order_id):
    """
    Allows a customer to request a refund within 3 days of delivery,
    choose from CEO-managed dynamic reasons, select specific order items with images and prices,
    and review the store policy/downloadable PDF.
    """
    order = get_object_or_404(StoreOrder, id=order_id, customer=request.user)
    
    # Verify order is delivered
    if order.status != 'delivered':
        messages.error(request, "Refund requests can only be initiated for delivered orders.")
        return redirect('store_order_detail', order_id=order.id)
        
    # Check 4-day policy window from updated_at or delivery time
    delivery_time = order.updated_at
    if timezone.now() > delivery_time + timedelta(days=4):
        messages.error(request, "The 4-day window for requesting a return or refund has expired.")
        return redirect('store_order_detail', order_id=order.id)
        
    # Check if a return request already exists for this order
    if StoreReturnRequest.objects.filter(order=order).exists():
        messages.warning(request, "A return request has already been submitted for this order.")
        return redirect('store_order_detail', order_id=order.id)

    # Fetch the official store return policy & dynamic reasons created by the CEO
    policy = StoreReturnPolicy.objects.first()
    refund_reasons = RefundReason.objects.filter(is_active=True)

    if request.method == 'POST':
        reason_id = request.POST.get('reason_id')
        issue_description = request.POST.get('issue_description')
        selected_item_ids = request.POST.getlist('selected_items') # List of StoreOrderItem IDs checked by customer
        
        # Updated to handle maximum 2 images, removed image_3 and video_proof
        image_1 = request.FILES.get('image_1')
        image_2 = request.FILES.get('image_2')

        if not selected_item_ids:
            messages.error(request, "Please select at least one item to return.")
            return redirect('request_store_refund', order_id=order.id)

        # Require at least 1 image, capped up to 2 images
        if not image_1:
            messages.error(request, "Please provide at least 1 image proof.")
            return redirect('request_store_refund', order_id=order.id)

        reason_obj = get_object_or_404(RefundReason, id=reason_id)

        # Create the return request container (storing up to 2 images)
        return_req = StoreReturnRequest.objects.create(
            order=order,
            customer=request.user,
            reason=reason_obj,
            issue_description=issue_description,
            image_1=image_1,
            image_2=image_2,
            status='pending_management_approval'
        )

        # Loop through selected order items and add them to the return request with their prices
        total_refund_calc = Decimal('0.00')
        for item_id in selected_item_ids:
            order_item = get_object_or_404(order.items, id=item_id)
            item_total = order_item.total_price
            total_refund_calc += item_total

            StoreReturnItem.objects.create(
                return_request=return_req,
                order_item=order_item,
                quantity_to_return=order_item.quantity,
                refund_item_total=item_total
            )

        return_req.refund_amount = total_refund_calc
        return_req.save()
            
        order.status = 'return_requested'
        order.save()
            
        ActivityAuditLog.objects.create(
            user=request.user,
            role='customer',
            action=f"Submitted Return/Refund request for Store Order #{order.id} with item selections."
        )
        messages.success(request, "Return and refund request submitted successfully with item details!")
        return redirect('store_order_detail', order_id=order.id)

    return render(request, 'store/request_refund.html', {
        'order': order, 
        'policy': policy, 
        'refund_reasons': refund_reasons
    })

@login_required
def management_return_dashboard_view(request):
    """Dashboard for CEO, General Manager, and Manager to review, inspect, assign riders, or reject returns."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') in ['CEO', 'General Manager', 'Manager']):
        messages.error(request, "Unauthorized access.")
        return redirect('store_home')

    returns_query = StoreReturnRequest.objects.all().order_by('-created_at')

    # Filtering options (Customer Name, Username, Amount, Date)
    customer_name = request.GET.get('customer_name')
    username = request.GET.get('username')
    amount = request.GET.get('amount')
    date_str = request.GET.get('date')

    if customer_name:
        returns_query = returns_query.filter(Q(customer__first_name__icontains=customer_name) | Q(customer__last_name__icontains=customer_name))
    if username:
        returns_query = returns_query.filter(customer__username__icontains=username)
    if amount:
        returns_query = returns_query.filter(refund_amount=amount)
    if date_str:
        returns_query = returns_query.filter(created_at__date=date_str)

    if request.method == 'POST':
        return_id = request.POST.get('return_id')
        action = request.POST.get('action')
        return_obj = get_object_or_404(StoreReturnRequest, id=return_id)

        if action == 'assign_rider':
            rider_id = request.POST.get('rider_id')
            rider = get_object_or_404(User, id=rider_id)
            return_obj.assigned_rider = rider
            return_obj.rider_status = 'going_to_site'
            return_obj.status = 'rider_assigned'
            return_obj.save()
            messages.success(request, f"Rider {rider.username} assigned for pickup successfully.")

        elif action == 'inspect_and_approve':
            return_obj.inspected_by = request.user
            return_obj.inspection_notes = request.POST.get('inspection_notes', '')
            return_obj.status = 'approved_pending_finance'
            return_obj.save()
            messages.success(request, "Return inspected and approved! Forwarded to Finance.")

        elif action == 'reject':
            return_obj.inspected_by = request.user
            return_obj.rejection_reason_note = request.POST.get('rejection_reason_note', '')
            if 'rejection_proof_file' in request.FILES:
                return_obj.rejection_proof_file = request.FILES['rejection_proof_file']
            return_obj.status = 'rejected_and_closed'
            return_obj.save()
            messages.warning(request, "Return request rejected. Customer notified to pick up item at store.")
            
        return redirect('management_return_dashboard') # <-- Updated to match your exact URL pattern name

    # Robust rider filtering to safely target riders regardless of profile setups
    riders = User.objects.filter(Q(role__iexact='rider') | Q(groups__name__iexact='rider')).distinct()
    if not riders.exists():
        riders = User.objects.filter(is_staff=True) # Fallback to staff if no specific rider role/group is found

    return render(request, 'store/management_returns.html', {
        'return_requests': returns_query,
        'riders': riders,
    })

@login_required
def rider_return_status_update_view(request, return_id):
    """Allows assigned riders to update status: Going to site -> Picked up -> Delivered to store."""
    return_obj = get_object_or_404(StoreReturnRequest, id=return_id, assigned_rider=request.user)
    
    if request.method == 'POST':
        new_rider_status = request.POST.get('rider_status')
        if new_rider_status in ['going_to_site', 'picked_up', 'delivered_to_store']:
            return_obj.rider_status = new_rider_status
            if new_rider_status == 'delivered_to_store':
                return_obj.status = 'delivered_to_store'
            else:
                return_obj.status = 'rider_assigned'
            return_obj.save()
            messages.success(request, f"Rider status updated to: {return_obj.get_rider_status_display()}")
            
    return redirect(request.META.get('HTTP_REFERER', 'store_home'))

from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from .models import StoreReturnRequest, ActivityAuditLog

@login_required
def finance_refund_settlement_view(request, return_id):
    """Finance view to confirm refund, input amount, attach invoice, and close request."""
    user_role = getattr(request.user, 'role', '')
    if not (request.user.is_superuser or user_role in ['store_finance', 'Finance', 'general_manager']):
        messages.error(request, "Unauthorized access.")
        return redirect('store_home')

    return_obj = get_object_or_404(StoreReturnRequest, id=return_id)

    if request.method == 'POST':
        try:
            raw_amount = request.POST.get('refund_amount', '0.00')
            refund_amount = Decimal(raw_amount)
        except (InvalidOperation, ValueError):
            messages.error(request, "Invalid refund amount format entered.")
            return redirect('store_finance_dashboard')

        refund_invoice = request.FILES.get('refund_invoice')
        
        return_obj.refund_amount = refund_amount
        if refund_invoice:
            return_obj.refund_invoice = refund_invoice
        return_obj.refund_processed_by = request.user
        return_obj.status = 'refunded_and_closed'
        return_obj.save()

        order = return_obj.order
        if hasattr(order, 'total_amount') and order.total_amount is not None:
            order.total_amount = max(Decimal('0.00'), order.total_amount - refund_amount)
        order.status = 'refunded'
        order.save()

        ActivityAuditLog.objects.create(
            user=request.user,
            role=user_role or 'store_finance',
            action=f"Processed refund of ₦{refund_amount} for Return Request #{return_obj.id}. Closed ticket."
        )
        messages.success(request, f"Refund settlement of ₦{refund_amount} completed successfully. Customer receipt generated.")
        
        return redirect('store_finance_dashboard')

    return render(request, 'store/finance_refund_settlement.html', {'return_obj': return_obj})

@login_required
def export_refunds_csv_view(request):
    """Allows Finance, CEO, and GM to export all refund history with timestamps and stages."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') in ['CEO', 'General Manager', 'Finance']):
        return HttpResponse("Unauthorized", status=403)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="refund_history_report.csv"'

    writer = csv.writer(response)
    writer.writerow(['Return ID', 'Order ID', 'Customer', 'Reason', 'Status', 'Rider Status', 'Refund Amount', 'Created At'])

    for req in StoreReturnRequest.objects.select_related('customer', 'reason', 'order'):
        writer.writerow([
            req.id,
            req.order.id,
            req.customer.username,
            req.reason.reason_text if req.reason else 'N/A',
            req.get_status_display(),
            req.get_rider_status_display(),
            req.refund_amount,
            req.created_at
        ])

    return response


@login_required
def ceo_manage_return_policy(request):
    """Allows CEO or Superuser to manage policy text/PDF and add/delete dynamic return reasons."""
    if not (request.user.is_superuser or getattr(request.user, 'role', '') in ['CEO', 'General Manager']):
        messages.error(request, "Unauthorized access.")
        return redirect('store_home')

    policy, created = StoreReturnPolicy.objects.get_or_create(id=1, defaults={
        'title': "TechsNi Store Return & Refund Policy",
        'content': "Items can be returned within 4 days of delivery with proper media proof..."
    })

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_policy':
            policy.title = request.POST.get('title', policy.title)
            policy.content = request.POST.get('content', policy.content)
            if 'policy_document' in request.FILES:
                policy.policy_document = request.FILES['policy_document']
            policy.updated_by = request.user
            policy.save()
            messages.success(request, "Return policy updated successfully and published to users.")

        elif action == 'add_reason':
            reason_text = request.POST.get('reason_text')
            if reason_text:
                RefundReason.objects.get_or_create(reason_text=reason_text)
                messages.success(request, "New return reason added successfully.")

        elif action == 'delete_reason':
            reason_id = request.POST.get('reason_id')
            RefundReason.objects.filter(id=reason_id).delete()
            messages.success(request, "Return reason removed successfully.")

        return redirect('ceo_manage_return_policy')

    refund_reasons = RefundReason.objects.all().order_by('-created_at')
    return render(request, 'store/ceo_return_policy_manager.html', {
        'policy': policy,
        'refund_reasons': refund_reasons
    })
